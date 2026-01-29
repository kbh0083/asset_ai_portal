"""
문서 파서 모듈

PDF 및 Excel 파일에서 텍스트를 추출하는 함수들을 제공합니다.
"""
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from langchain.messages import AIMessage

# docling imports
try:
    from docling.datamodel.base_models import InputFormat
    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling.datamodel.pipeline_options import PdfPipelineOptions, TableFormerMode
    from docling.datamodel.accelerator_options import AcceleratorOptions
    from docling.backend.docling_parse_v4_backend import DoclingParseV4DocumentBackend
    from docling_core.types.doc.document import ContentLayer
    DOCLING_AVAILABLE = True
except ImportError:
    # docling이 설치되지 않은 경우를 위한 처리
    DOCLING_AVAILABLE = False

# pikepdf import
try:
    import pikepdf
    PIKEPDF_AVAILABLE = True
except ImportError:
    PIKEPDF_AVAILABLE = False


def table_to_markdown(table: List[List]) -> str:
    """
    표 데이터를 마크다운 테이블 형식으로 변환하는 헬퍼 함수
    
    Args:
        table: 2차원 리스트 형태의 표 데이터
    
    Returns:
        마크다운 테이블 문자열
    """
    if not table:
        return ""
    
    # 빈 셀을 빈 문자열로 변환하고 파이프 문자 이스케이프 처리
    def clean_cell(cell):
        if cell is None:
            return ""
        cell_str = str(cell).strip()
        # 마크다운 테이블의 파이프 문자를 이스케이프
        cell_str = cell_str.replace("|", "\\|")
        return cell_str
    
    # 표 데이터 정리
    cleaned_table = [[clean_cell(cell) for cell in row] for row in table]
    
    # 최대 컬럼 수 확인
    max_cols = max(len(row) for row in cleaned_table) if cleaned_table else 0
    
    # 모든 행을 동일한 컬럼 수로 맞춤
    normalized_table = []
    for row in cleaned_table:
        normalized_row = row + [""] * (max_cols - len(row))
        normalized_table.append(normalized_row)
    
    if not normalized_table:
        return ""
    
    markdown_lines = []
    
    # 헤더 행 (첫 번째 행)
    header = normalized_table[0]
    markdown_lines.append("| " + " | ".join(header) + " |")
    
    # 구분선
    markdown_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
    
    # 데이터 행들
    for row in normalized_table[1:]:
        markdown_lines.append("| " + " | ".join(row) + " |")
    
    return "\n".join(markdown_lines)


def get_last_ai_message(test_result: Optional[Dict[str, Any]]) -> Optional[AIMessage]:
    """
    _test_result의 'messages' 배열에서 마지막 AIMessage를 가져와서 반환하는 함수
    
    Args:
        test_result: agent 실행 결과 딕셔너리 (messages 키 포함)
    
    Returns:
        마지막 AIMessage 객체 또는 None
    """
    if not test_result or not isinstance(test_result, dict) or "messages" not in test_result:
        return None
    
    messages = test_result["messages"]
    
    if not isinstance(messages, list):
        return None
    
    # 역순으로 순회하여 첫 번째 AIMessage 찾기
    for message in reversed(messages):
        if isinstance(message, AIMessage):
            return message
    
    return None


def get_file_path(file_path: str) -> str:
    """
    파일 경로를 절대 경로로 변환하는 함수
    
    Args:
        file_path: 상대 경로 또는 절대 경로
    
    Returns:
        절대 경로 문자열
    
    Raises:
        FileNotFoundError: 파일을 찾을 수 없을 때
    """
    # 파일 경로 확인 및 절대 경로로 변환
    file_path_obj = Path(file_path)
    if not file_path_obj.is_absolute():
        # 노트북 위치 기준 상대 경로 처리
        # 노트북은 asset_ai_portal/tests 폴더에 있고, documents는 20_code_test 루트에 있음
        current_dir = Path.cwd()
        
        # asset_ai_portal/tests에서 실행 중이면 상위로 두 번 이동 (20_code_test 루트)
        if current_dir.name == 'tests' and current_dir.parent.name == 'asset_ai_portal':
            project_root = current_dir.parent.parent  # tests -> asset_ai_portal -> 20_code_test
        elif current_dir.name == 'asset_ai_portal':
            project_root = current_dir.parent  # asset_ai_portal -> 20_code_test
        else:
            # 20_code_test에서 실행 중이면 그대로 사용
            project_root = current_dir
        
        file_path_obj = project_root / file_path_obj
    
    if not file_path_obj.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path_obj}")
    
    return str(file_path_obj)


def extract_pdf_with_pdfplumber(pdf_path: str, password: Optional[str] = None) -> str:
    """
    pdfplumber를 사용하여 PDF 파일을 마크다운 형식으로 변환하는 함수
    
    pdfplumber는 PDF 파일을 텍스트 데이터로 추출하는 라이브러리로, 표, 이미지, 레이아웃 등을 잘 보존합니다.
    암호화된 PDF와 암호화되지 않은 PDF 모두 처리할 수 있습니다.
    
    Args:
        pdf_path: PDF 파일 경로
        password: PDF 암호 (선택사항)
    
    Returns:
        마크다운 형식의 텍스트
    
    Raises:
        ImportError: pdfplumber가 설치되지 않았을 때
        ValueError: PDF 암호가 올바르지 않을 때
        Exception: 기타 PDF 처리 오류
    """
    try:
        import pdfplumber
    except ImportError as exc:
        raise ImportError(
            "PDF를 처리하기 위해 pdfplumber가 필요합니다.\n"
            "설치 명령: pip install pdfplumber"
        ) from exc
    
    markdown_parts = []
    
    try:
        pdf_path = get_file_path(pdf_path)
        # password가 있으면 암호화된 PDF로 처리, 없으면 암호화되지 않은 PDF로 처리
        pdf_kwargs = {"password": password} if password else {}
        
        with pdfplumber.open(pdf_path, **pdf_kwargs) as pdf:
            for page in pdf.pages:
                page_content = []
                
                # 표 추출 (표가 있으면 먼저 표를 추출)
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if table:
                            markdown_table = table_to_markdown(table)
                            if markdown_table:
                                page_content.append(markdown_table)
                                page_content.append("")  # 표 다음에 빈 줄 추가
                
                # 텍스트 추출
                text = page.extract_text()
                if text:
                    page_content.append(text)
                
                if page_content:
                    markdown_parts.append("\n".join(page_content))
        
        return "\n\n".join(markdown_parts) if markdown_parts else ""
        
    except Exception as e:
        # 암호화 관련 오류인지 확인
        error_msg = str(e).lower()
        if 'password' in error_msg or 'encrypted' in error_msg or 'incorrect password' in error_msg:
            raise ValueError(f"PDF 암호가 올바르지 않거나 암호화된 PDF를 읽을 수 없습니다: {e}") from e
        raise


def _decrypt_to_temp_pdf(src_pdf: str, password: Optional[str]) -> Path:
    """
    pikepdf로 PDF를 열어 복호화된 PDF를 임시 파일로 저장 후 경로 반환
    
    Args:
        src_pdf: 원본 PDF 파일 경로
        password: PDF 암호 (None이면 암호화되지 않은 PDF로 처리)
    
    Returns:
        복호화된 임시 PDF 파일의 Path 객체
    
    Raises:
        ImportError: pikepdf가 설치되지 않았을 때
        ValueError: PDF 복호화 실패 시
    """
    if not PIKEPDF_AVAILABLE:
        raise ImportError(
            "암호화된 PDF를 처리하기 위해 pikepdf가 필요합니다.\n"
            "설치 명령: pip install pikepdf"
        )
    
    src_pdf_path = Path(src_pdf)
    
    if not src_pdf_path.exists():
        raise FileNotFoundError(f"PDF 파일을 찾을 수 없습니다: {src_pdf_path}")
    
    # 임시 파일 생성
    temp_dir = tempfile.gettempdir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_filename = f"{src_pdf_path.stem}_{timestamp}{src_pdf_path.suffix}"
    decrypted_pdf = Path(temp_dir) / temp_filename

    try:
        if password:
            with pikepdf.open(str(src_pdf_path), password=password) as pdf:
                pdf.save(str(decrypted_pdf))
        else:
            # 암호화되지 않은 PDF도 복사하여 일관성 유지
            with pikepdf.open(str(src_pdf_path)) as pdf:
                pdf.save(str(decrypted_pdf))
        return decrypted_pdf
    except Exception as e:
        # 실패 시 임시 파일이 생성되었을 수 있으므로 정리
        if decrypted_pdf.exists():
            try:
                decrypted_pdf.unlink()
            except OSError:
                pass
        raise ValueError(f"PDF 복호화 실패: {e}") from e


def extract_text_from_pdf_with_docling_nopassword(pdf_path: str) -> str:
    """
    docling을 사용하여 암호화되지 않은 PDF 파일에서 텍스트를 추출하는 함수
    
    Args:
        pdf_path: PDF 파일 경로
    
    Returns:
        마크다운 형식의 텍스트
    
    Raises:
        ImportError: docling이 설치되지 않았을 때
        RuntimeError: PDF 처리 중 오류 발생 시
    """
    if not DOCLING_AVAILABLE:
        raise ImportError(
            "PDF를 처리하기 위해 docling이 필요합니다.\n"
            "설치 명령: pip install docling"
        )
    
    pdf_path = get_file_path(pdf_path)

    def _run(do_cell_matching: bool) -> str:
        """내부 실행 함수"""
        pipeline = PdfPipelineOptions(
            do_ocr=False,
            do_table_structure=True,
            do_picture_classification=False,
            do_picture_description=False,
            generate_page_images=False,
            images_scale=2.0,  # 레이아웃/테이블 크롭 품질에 도움될 수 있음
        )

        # 표 구조 품질 우선
        pipeline.table_structure_options.mode = TableFormerMode.ACCURATE
        pipeline.table_structure_options.do_cell_matching = do_cell_matching

        # 표 구조 목적이면 force_backend_text는 끄는 쪽이 안전
        pipeline.force_backend_text = False

        pipeline.accelerator_options = AcceleratorOptions(device="cpu")

        converter = DocumentConverter(
            allowed_formats=[InputFormat.PDF],
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_options=pipeline,
                    backend=DoclingParseV4DocumentBackend,
                )
            },
        )

        result = converter.convert(pdf_path)
        md = result.document.export_to_markdown(
            included_content_layers={ContentLayer.BODY},
        )

        return md

    try:
        md = _run(do_cell_matching=True)
        return md
    except Exception as e:
        # 예외를 다시 발생시켜 호출자가 처리할 수 있도록 함
        raise RuntimeError(f"PDF 텍스트 추출 실패: {e}") from e


def extract_pdf_with_docling(pdf_path: str, password: Optional[str] = None) -> str:
    """
    docling을 사용하여 PDF 파일에서 텍스트를 추출하는 함수
    
    Args:
        pdf_path: PDF 파일 경로
        password: PDF 암호 (선택사항)
    
    Returns:
        마크다운 형식의 텍스트
    
    Raises:
        ImportError: docling 또는 pikepdf가 설치되지 않았을 때
        RuntimeError: PDF 처리 중 오류 발생 시
    """
    if password:
        # 암호화된 PDF는 임시 파일로 복호화 후 처리
        temp_pdf_path = None
        try:
            temp_pdf_path = _decrypt_to_temp_pdf(pdf_path, password)
            pdf_text = extract_text_from_pdf_with_docling_nopassword(str(temp_pdf_path))
            return pdf_text
        finally:
            # 임시 파일 삭제
            if temp_pdf_path and temp_pdf_path.exists():
                try:
                    temp_pdf_path.unlink()
                except OSError:
                    # 삭제 실패해도 계속 진행 (임시 파일이므로)
                    pass
    else:
        return extract_text_from_pdf_with_docling_nopassword(pdf_path)


def extract_text_from_excel(excel_path: str) -> Dict[str, str]:
    """
    Excel 파일(.xlsx, .xls)에서 모든 시트의 텍스트를 추출하는 함수
    
    Args:
        excel_path: Excel 파일 경로 (상대 경로 또는 절대 경로)
    
    Returns:
        시트 이름을 키로 하고 추출된 텍스트를 값으로 하는 딕셔너리
    
    Raises:
        FileNotFoundError: Excel 파일을 찾을 수 없을 때
        ImportError: 필요한 Excel 라이브러리가 설치되지 않았을 때
    """
    # 파일 경로 확인 및 절대 경로로 변환    
    excel_path = get_file_path(excel_path)

    print(f"excel_path: {excel_path}")  
    
    # 파일 확장자 확인 (Path 객체로 변환)
    excel_path_obj = Path(excel_path)
    file_ext = excel_path_obj.suffix.lower()
    
    # 여러 Excel 라이브러리 시도 (우선순위 순)
    # 1. pandas + openpyxl/xlrd (가장 편리함)
    try:
        import pandas as pd
        
        # 모든 시트 읽기
        if file_ext == '.xlsx':
            excel_file = pd.ExcelFile(str(excel_path), engine='openpyxl')
        elif file_ext == '.xls':
            excel_file = pd.ExcelFile(str(excel_path), engine='xlrd')
        else:
            # 자동 감지
            excel_file = pd.ExcelFile(str(excel_path))
        
        sheets_text = {}
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            # DataFrame을 텍스트로 변환
            text_parts = []
            # 헤더 포함하여 모든 셀의 값을 문자열로 변환
            for _, row in df.iterrows():
                row_values = [str(val) if pd.notna(val) else '' for val in row.values]
                text_parts.append(' | '.join(row_values))
            
            sheets_text[sheet_name] = '\n'.join(text_parts)

        print("using pandas")
        return sheets_text
    except ImportError as e:
        if 'pandas' in str(e):
            pass  # pandas가 없으면 다음 방법 시도
        elif 'openpyxl' in str(e) or 'xlrd' in str(e):
            # pandas는 있지만 엔진이 없는 경우
            raise ImportError(
                "Excel 파일을 읽기 위한 엔진이 필요합니다.\n"
                ".xlsx 파일: pip install openpyxl\n"
                ".xls 파일: pip install xlrd"
            ) from e
        else:
            raise
    
    # 2. openpyxl (xlsx 파일용)
    if file_ext == '.xlsx':
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(str(excel_path), data_only=True)
            sheets_text = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts = []
                
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(val) if val is not None else '' for val in row]
                    text_parts.append(' | '.join(row_values))
                
                sheets_text[sheet_name] = '\n'.join(text_parts)
            
            print("using openpyxl")
            return sheets_text
        except ImportError:
            pass
    
    # 3. xlrd (xls 파일용)
    if file_ext == '.xls':
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(str(excel_path))
            sheets_text = {}
            
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                text_parts = []
                
                for row_idx in range(sheet.nrows):
                    row_values = [str(sheet.cell_value(row_idx, col_idx)) 
                                 for col_idx in range(sheet.ncols)]
                    text_parts.append(' | '.join(row_values))
                
                sheets_text[sheet_name] = '\n'.join(text_parts)
            
            print("using xlrd")
            return sheets_text
        except ImportError:
            pass
    
    # 모든 라이브러리가 없으면 에러
    raise ImportError(
        "Excel 텍스트 추출을 위한 라이브러리가 설치되지 않았습니다.\n"
        "다음 중 하나를 설치해주세요:\n"
        "  - pandas + openpyxl (권장): pip install pandas openpyxl\n"
        "  - pandas + xlrd (.xls 파일용): pip install pandas xlrd\n"
        "  - openpyxl (.xlsx 파일용): pip install openpyxl\n"
        "  - xlrd (.xls 파일용): pip install xlrd"
    )


def parser_excel(excel_path: str) -> str:
    """
    Excel 파일의 모든 시트 텍스트를 하나의 문자열로 반환하는 편의 함수
    
    Args:
        excel_path: Excel 파일 경로
    
    Returns:
        모든 시트의 텍스트를 합친 문자열
    """
    sheets_dict = extract_text_from_excel(excel_path)
    
    result_parts = []
    for sheet_name, sheet_text in sheets_dict.items():
        result_parts.append(f"=== 시트: {sheet_name} ===")
        result_parts.append(sheet_text)
        result_parts.append("")  # 빈 줄 추가
    
    return '\n'.join(result_parts)
